import re
import json
import requests
import logging
from django.db import transaction
from django.db.models import Q, Sum
from datetime import datetime, time
from .models import ScheduleSlot, Subject, Classroom, TimeSlot, PlanDiscipline, SubjectTemplate, AcademicPlan
from .utils import safe_int, safe_float, safe_str        
from accounts.models import Group, Teacher, User, Department
import openpyxl

logger = logging.getLogger(__name__)


class ScheduleImporter:
    DAYS_MAP = {
        'душанбе': 0, 'понедельник': 0, 'monday': 0,
        'сешанбе': 1, 'вторник': 1, 'tuesday': 1,
        'чоршанбе': 2, 'среда': 2, 'wednesday': 2,
        'панҷшанбе': 3, 'четверг': 3, 'thursday': 3,
        'ҷумъа': 4, 'пятница': 4, 'friday': 4,
        'шанбе': 5, 'суббота': 5, 'saturday': 5,
    }

    def __init__(self, file=None):
        self.file = file
        self.preview_data = []
        self.default_dept = Department.objects.first()
        self.all_groups = list(Group.objects.all())

    def parse_for_preview(self, default_group=None):
        filename = self.file.name.lower()
        self.default_group = default_group

        logger.info(
            "ScheduleImporter.parse_for_preview: file=%s default_group=%s",
            filename, default_group
        )

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                self._process_excel()
            else:
                logger.warning(
                    "ScheduleImporter.parse_for_preview: unsupported file format=%s", filename
                )
                return []
        except Exception:
            logger.exception("ScheduleImporter.parse_for_preview: failed to parse file=%s", filename)
            return []

        logger.info(
            "ScheduleImporter.parse_for_preview: parsed %s entries from %s",
            len(self.preview_data), filename
        )
        return self.preview_data

    def _process_excel(self):
        logger.debug("ScheduleImporter._process_excel: loading workbook")
        wb = openpyxl.load_workbook(self.file, data_only=True)
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            clean = [str(c).strip() if c is not None else "" for c in row]
            if any(clean):
                rows.append(clean)
        logger.debug("ScheduleImporter._process_excel: loaded %s non-empty rows", len(rows))
        self._parse_grid_rows(rows, source="Excel")

    def _parse_grid_rows(self, rows, source=""):
        if len(rows) < 2:
            logger.warning("ScheduleImporter._parse_grid_rows: too few rows (%s) in source=%s", len(rows), source)
            return

        header_map = {}
        start_row_idx = 0

        for row_idx, row in enumerate(rows[:20]):
            for col_idx, cell_val in enumerate(row):
                if not cell_val:
                    continue
                group = self._match_group(cell_val)
                if group:
                    header_map[col_idx] = group
            if header_map:
                start_row_idx = row_idx + 1
                break

        if not header_map and self.default_group:
            logger.debug(
                "ScheduleImporter._parse_grid_rows: no header groups found, using default_group=%s",
                self.default_group
            )
            header_map[2] = self.default_group
            start_row_idx = 0

        if not header_map:
            logger.warning(
                "ScheduleImporter._parse_grid_rows: no group headers found and no default_group set"
            )

        logger.debug(
            "ScheduleImporter._parse_grid_rows: found %s group columns, start_row=%s",
            len(header_map), start_row_idx
        )

        current_day = 0

        for row in rows[start_row_idx:]:
            row_text = " ".join([str(x).lower() for x in row if x])

            for d_name, d_val in self.DAYS_MAP.items():
                if d_name in row_text[:30]:
                    current_day = d_val
                    break

            is_military_row = 'ҳарбӣ' in row_text or 'военная' in row_text

            current_time_slot = None
            for c_val in row[:4]:
                ts = self._extract_time_slot(c_val)
                if ts:
                    current_time_slot = ts
                    break

            if not current_time_slot and not is_military_row:
                continue

            for col_idx, group in header_map.items():
                if col_idx >= len(row):
                    continue

                raw_val = row[col_idx]
                if not raw_val and not is_military_row:
                    continue
                if raw_val and len(raw_val) < 2 and not is_military_row:
                    continue

                room = ""
                if col_idx + 1 < len(row):
                    pot_room = row[col_idx + 1]
                    if pot_room and len(pot_room) < 10 and any(c.isdigit() for c in pot_room):
                        room = pot_room

                if is_military_row:
                    if not current_time_slot:
                        times = ["08:00-08:50", "09:00-09:50", "10:00-10:50"]
                        for t in times:
                            self.preview_data.append({
                                'group_id': group.id,
                                'group_name': group.name,
                                'day': current_day,
                                'start_time': t.split('-')[0],
                                'end_time': t.split('-')[1],
                                'subject': "Кафедраи ҳарбӣ",
                                'teacher': "",
                                'room': "",
                                'type': "PRACTICE",
                                'is_military': True
                            })
                        continue
                    else:
                        raw_val = "Кафедраи ҳарбӣ"

                parsed_data = self._parse_cell_text(raw_val)

                self.preview_data.append({
                    'group_id': group.id,
                    'group_name': group.name,
                    'day': current_day,
                    'start_time': current_time_slot.start_time.strftime("%H:%M"),
                    'end_time': current_time_slot.end_time.strftime("%H:%M"),
                    'subject': parsed_data['subject'],
                    'teacher': parsed_data['teacher'],
                    'room': room,
                    'type': parsed_data['type'],
                    'is_military': is_military_row
                })

    def save_data_from_preview(self, clean_data, semester):
        stats = {'slots': 0, 'subjects': 0, 'teachers': 0}

        logger.info(
            "ScheduleImporter.save_data_from_preview: saving %s items for semester=%s",
            len(clean_data), semester
        )

        for item in clean_data:
            subj_name = safe_str(item.get('subject', '')).strip()
            if not subj_name:
                logger.debug("ScheduleImporter.save_data_from_preview: empty subject_name, skip item=%s", item)
                continue

            try:
                with transaction.atomic():
                    group = Group.objects.get(id=item['group_id'])

                    institute = None
                    if group.specialty and group.specialty.department.faculty:
                        institute = group.specialty.department.faculty.institute

                    start_str = item['start_time']
                    t_slot = None

                    if institute:
                        t_slot = TimeSlot.objects.filter(
                            institute=institute,
                            start_time__startswith=start_str[:5]
                        ).first()

                    if not t_slot:
                        t_slot = TimeSlot.objects.filter(
                            start_time__startswith=start_str[:5]
                        ).first()

                    if not t_slot:
                        logger.warning(
                            "ScheduleImporter.save_data_from_preview: no TimeSlot for start_time=%s "
                            "group=%s subject=%s — skipped",
                            start_str, group.name, subj_name
                        )
                        continue

                    dept = group.specialty.department if group.specialty else self.default_dept

                    is_mil = item.get('is_military') == 'on' or subj_name == 'Кафедраи ҳарбӣ'

                    subject, created = Subject.objects.get_or_create(
                        name__iexact=subj_name,
                        defaults={
                            'name': subj_name,
                            'code': f"IMP{hash(subj_name)}",
                            'department': dept,
                            'type': item.get('type', 'LECTURE'),
                            'is_stream_subject': False
                        }
                    )
                    if created:
                        stats['subjects'] += 1
                        logger.info(
                            "ScheduleImporter: created new Subject=%s dept=%s",
                            subj_name, dept
                        )
                    subject.groups.add(group)

                    teacher_obj = None
                    t_name = safe_str(item.get('teacher', '')).strip()
                    if t_name and len(t_name) > 2 and not is_mil:
                        tsplit = t_name.split()
                        surname = tsplit[0]
                        teacher_qs = Teacher.objects.filter(
                            Q(user__last_name__icontains=surname) |
                            Q(user__first_name__icontains=surname)
                        )
                        if teacher_qs.exists():
                            teacher_obj = teacher_qs.first()
                            logger.debug(
                                "ScheduleImporter: matched teacher=%s for name=%s",
                                teacher_obj.user.get_full_name(), t_name
                            )
                        else:
                            try:
                                base_u = f"imp_{datetime.now().microsecond}_{surname[:5]}"
                                u = User.objects.create_user(
                                    username=base_u,
                                    password='password123',
                                    role='TEACHER',
                                    last_name=surname,
                                    first_name=t_name[:100]
                                )
                                teacher_obj = Teacher.objects.create(user=u, department=dept)
                                stats['teachers'] += 1
                                logger.info(
                                    "ScheduleImporter: created new Teacher=%s username=%s",
                                    t_name, base_u
                                )
                            except Exception as te:
                                logger.warning(
                                    "ScheduleImporter: failed to create teacher name=%s: %s",
                                    t_name, te
                                )

                    if teacher_obj and not subject.teacher:
                        subject.teacher = teacher_obj
                        subject.save()

                    room_obj = None
                    room_text = safe_str(item.get('room', '')).strip()
                    if room_text and institute:
                        building = institute.buildings.first()
                        if building:
                            room_obj, _ = Classroom.objects.get_or_create(
                                building=building,
                                number=room_text,
                                defaults={'floor': 1}
                            )

                    ScheduleSlot.objects.filter(
                        group=group, semester=semester,
                        day_of_week=item['day'], time_slot=t_slot
                    ).delete()

                    ScheduleSlot.objects.create(
                        group=group,
                        semester=semester,
                        day_of_week=item['day'],
                        time_slot=t_slot,
                        subject=subject,
                        teacher=teacher_obj if not is_mil else None,
                        classroom=room_obj,
                        room=room_text,
                        lesson_type=item.get('type', 'LECTURE'),
                        is_military=is_mil,
                        is_active=True,
                        start_time=t_slot.start_time,
                        end_time=t_slot.end_time
                    )
                    stats['slots'] += 1

            except Group.DoesNotExist:
                logger.error(
                    "ScheduleImporter.save_data_from_preview: Group id=%s not found",
                    item.get('group_id')
                )
            except Exception as e:
                logger.warning(
                    "ScheduleImporter.save_data_from_preview: failed to save slot "
                    "group_id=%s subject=%s day=%s time=%s: %s",
                    item.get('group_id'), subj_name,
                    item.get('day'), item.get('start_time'), e
                )
                continue

        logger.info(
            "ScheduleImporter.save_data_from_preview finished: slots=%s subjects=%s teachers=%s",
            stats['slots'], stats['subjects'], stats['teachers']
        )
        return stats

    def _extract_time_slot(self, val):
        clean = re.sub(r'[^\d]', '', str(val))
        if len(clean) >= 4:
            try:
                h, m = int(clean[:2]), int(clean[2:4])
                return TimeSlot.objects.filter(
                    start_time__hour=h, start_time__minute=m
                ).first()
            except Exception:
                pass
        return None

    def _match_group(self, text):
        t = str(text).lower().replace(" ", "")
        for g in self.all_groups:
            if g.name.lower().replace(" ", "") in t:
                return g
            if re.search(r'\d-\d{5,}', t):
                if t.startswith(g.name.lower()[:5]):
                    return g
        return None

    def _parse_cell_text(self, text):
        text = str(text).replace('\n', ' ').strip()

        l_type = 'LECTURE'
        low = text.lower()
        if any(x in low for x in ['(а)', '(амалӣ)', 'pr']):
            l_type = 'PRACTICE'
        elif any(x in low for x in ['(к)', '(кмро)', 'srsp']):
            l_type = 'SRSP'

        subj = text
        teach = ""

        titles = ['Дот', 'Асс', 'Проф', 'Омузгор', 'Муаллим']

        for title in titles:
            if title in text:
                parts = text.split(title)
                if len(parts) > 1:
                    subj = parts[0]
                    teach = title + parts[1]
                break

        if not teach and ')' in text:
            parts = text.split(')')
            subj = parts[0] + ')'
            if len(parts) > 1 and len(parts[1].strip()) > 3:
                teach = parts[1]

        subj = re.sub(r'\(.*?\)', '', subj).strip()
        if not subj:
            subj = "Нераспознанный предмет"

        teach = re.sub(r'[^\w\s\.]', '', teach).strip()

        return {'subject': subj, 'teacher': teach, 'type': l_type}


class RupImporter:
    @staticmethod
    def parse_for_preview(file):
        logger.info("RupImporter.parse_for_preview: starting parse")
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active

        preview_data = []
        current_disc_type = 'REQUIRED'

        _REQUIRED_MARKERS = {'ҳатмӣ', 'обязательн'}
        _ELECTIVE_MARKERS  = {'интихобӣ', 'ихтиёрӣ', 'выбор'}
        _SKIP_MARKERS      = {'номгӯйи', 'наименование', 'семестр', 'итого', 'барча', 'всего'}

        from schedule.models import SubjectTemplate

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            raw_name = row[0] if row else None
            subj_name = safe_str(raw_name).strip()

            if not subj_name:
                continue

            lower_name = subj_name.lower()

            if any(m in lower_name for m in _REQUIRED_MARKERS):
                current_disc_type = 'REQUIRED'
                logger.debug("RupImporter: row=%s — switching to REQUIRED block", row_idx)
                continue
            if any(m in lower_name for m in _ELECTIVE_MARKERS):
                current_disc_type = 'ELECTIVE'
                logger.debug("RupImporter: row=%s — switching to ELECTIVE block", row_idx)
                continue
            if any(m in lower_name for m in _SKIP_MARKERS):
                logger.debug("RupImporter: row=%s — skip marker found: %s", row_idx, subj_name)
                continue

            credits = safe_int(row[1]) if len(row) > 1 else 0
            lec     = safe_int(row[4]) if len(row) > 4 else 0
            prac    = safe_int(row[5]) if len(row) > 5 else 0
            srsp    = safe_int(row[6]) if len(row) > 6 else 0
            srs     = safe_int(row[7]) if len(row) > 7 else 0

            if lec == 0 and prac == 0 and srsp == 0 and srs == 0:
                total_h = safe_int(row[3]) if len(row) > 3 else 0
                if total_h > 0:
                    srs = total_h
                    logger.debug(
                        "RupImporter: row=%s name=%s — no hour breakdown, using total=%s as SRS",
                        row_idx, subj_name, total_h
                    )

            is_new = not SubjectTemplate.objects.filter(name__iexact=subj_name).exists()

            preview_data.append({
                'id':      row_idx,
                'name':    subj_name,
                'type':    current_disc_type,
                'credits': credits,
                'lec':     lec,
                'prac':    prac,
                'srsp':    srsp,
                'srs':     srs,
                'is_new_template': is_new,
            })

        logger.info(
            "RupImporter.parse_for_preview: found %s disciplines", len(preview_data)
        )
        return preview_data


class AlgorithmicAssignmentService:
    @staticmethod
    def generate_assignment(teachers_qs, subjects_data):
        logger.info(
            "AlgorithmicAssignmentService.generate_assignment: teachers=%s subjects=%s",
            teachers_qs.count(), len(subjects_data)
        )
        assignments = []

        teacher_loads = {}
        for t in teachers_qs:
            load = t.subject_set.aggregate(
                total=Sum('lecture_hours') + Sum('practice_hours') + Sum('control_hours')
            )['total']
            teacher_loads[t.id] = load if load else 0

        for subj in subjects_data:
            key = subj['key']
            subj_name_lower = subj['name'].lower()
            subj_words = set(re.findall(r'\b\w{4,}\b', subj_name_lower))

            try:
                disc_id = int(key.split('_')[1])
                disc = PlanDiscipline.objects.get(id=disc_id)
                target_dept = disc.plan.specialty.department if disc.plan.specialty else None
                disc_hours = disc.lecture_hours + disc.practice_hours + disc.control_hours
            except Exception as e:
                logger.warning(
                    "AlgorithmicAssignmentService: failed to resolve discipline key=%s: %s",
                    key, e
                )
                disc = None
                target_dept = None
                disc_hours = 30

            best_teacher = None
            best_score = -9999
            reason_text = ""

            for t in teachers_qs:
                score = 0
                match_reasons = []

                if target_dept and (
                    t.department == target_dept or
                    target_dept in t.additional_departments.all()
                ):
                    score += 50
                    match_reasons.append("Своя кафедра")

                for comp in t.competencies.all():
                    comp_name_lower = comp.name.lower()
                    comp_words = set(re.findall(r'\b\w{4,}\b', comp_name_lower))

                    if comp_name_lower in subj_name_lower or subj_name_lower in comp_name_lower:
                        score += 40
                        match_reasons.append(f"Компетенция '{comp.name}'")
                    elif len(subj_words.intersection(comp_words)) > 0:
                        score += 20
                        match_reasons.append(f"Тег '{comp.name}'")

                current_load = teacher_loads.get(t.id, 0)
                score -= (current_load * 0.5)

                if score > best_score:
                    best_score = score
                    best_teacher = t
                    reason_text = (
                        f"Алгоритм: {', '.join(match_reasons)}. "
                        f"Нагрузка: {current_load}ч."
                    )

            if best_teacher and best_score > -100:
                assignments.append({
                    "key": key,
                    "teacher_id": best_teacher.id,
                    "reason": reason_text
                })
                teacher_loads[best_teacher.id] += disc_hours
                logger.debug(
                    "AlgorithmicAssignmentService: assigned subject=%s to teacher=%s score=%.1f reason=%s",
                    subj['name'], best_teacher.user.get_full_name(), best_score, reason_text
                )
            else:
                assignments.append({
                    "key": key,
                    "teacher_id": None,
                    "reason": "Алгоритм: Подходящий преподаватель не найден"
                })
                logger.warning(
                    "AlgorithmicAssignmentService: no suitable teacher for subject=%s best_score=%.1f",
                    subj['name'], best_score
                )

        logger.info(
            "AlgorithmicAssignmentService.generate_assignment finished: assigned=%s unassigned=%s",
            sum(1 for a in assignments if a['teacher_id']),
            sum(1 for a in assignments if not a['teacher_id'])
        )
        return {"assignments": assignments}

class AIAssignmentService:
    OLLAMA_URL = "http://localhost:11434/api/generate"

    @staticmethod
    def generate_assignment(teachers, subjects, model_name="gemma3:4b"):
        teachers_text = ""
        for t in teachers:
            competencies = ", ".join([c.name for c in t.competencies.all()])
            teachers_text += (
                f"ID={t.id} | ФИО: {t.user.get_full_name()} | "
                f"Кафедра: {t.department.name if t.department else 'Нет'} | "
                f"Компетенции: {competencies} | "
                f"Интересы: {t.research_interests}\n"
            )

        subjects_text = ""
        for s in subjects:
            subjects_text += f"Key={s['key']} | Предмет: {s['name']}\n"

        prompt = f"""Ты — ИИ-ассистент учебной части университета.
Твоя задача: распределить дисциплины между доступными преподавателями на основе их кафедры, компетенций и научных интересов.
Учитывай таджикские и русские названия. Математик не должен вести философию.
Если ни один преподаватель не подходит по профилю, обязательно укажи teacher_id: null.

ДОСТУПНЫЕ ПРЕПОДАВАТЕЛИ:
{teachers_text}

ПРЕДМЕТЫ ДЛЯ РАСПРЕДЕЛЕНИЯ:
{subjects_text}

Верни ТОЛЬКО чистый валидный JSON без тегов markdown. Формат строго такой:
{{
  "assignments":[
    {{
      "key": "значение Key предмета",
      "teacher_id": 123,
      "reason": "Краткая причина выбора"
    }}
  ]
}}"""

        payload = {
            "model": model_name,
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": 0.1,
                "num_predict": 2048
            }
        }

        try:
            resp = requests.post(AIAssignmentService.OLLAMA_URL, json=payload, timeout=900)
            if resp.status_code == 200:
                raw_response = resp.json().get("response", "")
                return AIAssignmentService.extract_json(raw_response)
            else:
                logger.error(f"Ollama API Error: HTTP {resp.status_code}")
        except Exception as e:
            logger.error(f"AIAssignmentService Exception: {str(e)}")

        return None

    @staticmethod
    def extract_json(raw: str) -> dict:
        match = re.search(r'\{[\s\S]*\}', raw)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
        return {"assignments": []}